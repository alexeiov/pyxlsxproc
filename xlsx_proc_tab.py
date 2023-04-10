"""this module uses cell tab as mark"""
# но сработало и без модификации под отступы, только с цветом, с интесмо, только немного надо эксель руками отредактировать, что быстрее, чем править код.
import openpyxl as opx
import datetime
from pathlib import Path
import config

open_path = input('Enter path to file:')
file_to_process = input('Enter filename:')
result_file = file_to_process[:-5] + '_to_db_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
result_file_1 = file_to_process[:-5] + '_to_db_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + '.txt'

full_data_path = Path(open_path).joinpath(file_to_process)
full_save_path = Path(open_path).joinpath(result_file)
full_save_path_1 = Path(open_path).joinpath(result_file_1)


def process_xlsx(data_open_path, data_workbook_num=0):
    wb = opx.load_workbook(filename=data_open_path)

    """Default location of the data to be processed - first sheet of Excel workbook"""

    sheet = wb.worksheets[data_workbook_num]
    print('xlsx data read completed\n')

    row_num = sheet.max_row
    col_num = sheet.max_column

    data_l = [[None for t in range(config.original_columns_num)] for k in range(row_num)]

    """Variables below to be changed dependant on the structure of the data processed"""

    client_class_name_prev = None
    subdiv_name_prev = None
    # name_prev = None

    for r_num in range(1, row_num+1):
        data_l[r_num - 1][0] = config.div
        """Color index should be checked prior to using it in next condition: could be decimal or hex 
        in different files"""
        # print(sheet.cell(r_num, 2).fill.start_color.index)
        if sheet.cell(r_num, 2).value is None and sheet.cell(r_num, 2).fill.start_color.index == 28:  # 'FFF5F2DD'
            subdiv_name_prev = str(sheet.cell(r_num, 1).value).strip()
            # data_l[r_num - 1][1] = str(sheet.cell(r_num, 1).value).strip()
            data_l[r_num - 1][1] = subdiv_name_prev
            # client_class_name_prev = None
            # name_prev = None
            # inv_prev = None
            # mol_name_prev_prev = None
            # date_in_prev = None
            # okof_prev = None
            # status = None
            # gbv_prev = None
            # nbv_prev = None
            # comments_prev = None
        else:
            data_l[r_num - 1][1] = subdiv_name_prev

        if sheet.cell(r_num, 2).value is None and sheet.cell(r_num, 2).fill.start_color.index == 9:  # 'FFFFFFFF'
            client_class_name_prev = str(sheet.cell(r_num, 1).value).strip()
            data_l[r_num - 1][2] = str(sheet.cell(r_num, 1).value).strip()
            # name_prev = None
            # inv_prev = None
            # mol_name_prev_prev = None
            # date_in_prev = None
            # okof_prev = None
            # status = None
            # gbv_prev = None
            # nbv_prev = None
            # comments_prev = None
        else:
            data_l[r_num - 1][2] = client_class_name_prev

        if sheet.cell(r_num, 2).value is not None:
            data_l[r_num - 1][3] = sheet.cell(r_num, 1).value.strip()  # name
            '''Keeping original inventory number format with leading zeros'''
            data_l[r_num - 1][4] = '0'
            for i in range(9 - 1 - len(str(sheet.cell(r_num, 2).value).strip())):
                data_l[r_num - 1][4] += '0'
            data_l[r_num - 1][4] += str(sheet.cell(r_num, 2).value).strip() # inv

            # data_l[r_num - 1][5] = str(sheet.cell(r_num, 3).value).strip()  # OKOF
            # data_l[r_num - 1][6] = str(sheet.cell(r_num, 6).value).strip()  # subdiv_name
            # data_l[r_num - 1][7] = str(sheet.cell(r_num, 7).value).strip()  # subdiv
            # data_l[r_num - 1][8] = str(sheet.cell(r_num, 8).value).strip()  # HC
            data_l[r_num - 1][5] = str(sheet.cell(r_num, 3).value).strip()  # date_in
            # data_l[r_num - 1][10] = str(sheet.cell(r_num, 10).value).strip()  # HC_sign
            # data_l[r_num - 1][11] = str(sheet.cell(r_num, 11).value).strip()  # currency
            # data_l[r_num - 1][12] = str(sheet.cell(r_num, 12).value).strip()  # OKOF_OLD
            # data_l[r_num - 1][13] = str(sheet.cell(r_num, 13).value).strip()  # CL_CLASS_NAME
            data_l[r_num - 1][6] = str(sheet.cell(r_num, 4).value).strip()  # titul
            data_l[r_num - 1][7] = str(sheet.cell(r_num, 5).value).strip()  # GBV
            # data_l[r_num - 1][16] = str(sheet.cell(r_num, 16).value).strip()  # acc_dep
            data_l[r_num - 1][8] = str(sheet.cell(r_num, 7).value).strip()  # NBV
            data_l[r_num - 1][9] = str(sheet.cell(r_num, 9).value).strip()
            data_l[r_num - 1][10] = str(sheet.cell(r_num, 10).value).strip()

    return data_l


def save_results_to_xlsx(results, save_path):
    wb = opx.Workbook()
    sheet = wb.worksheets[0]
    sheet.title = 'data_to_db'
    sheet['A1'] = 'div'
    sheet['B1'] = 'subdiv_name'
    sheet['C1'] = 'client_cl_name'
    sheet['D1'] = 'name'
    sheet['E1'] = 'inv'
    sheet['F1'] = 'date_in'
    sheet['G1'] = 'TITUL'
    sheet['H1'] = 'OKOF'
    sheet['I1'] = 'sign'
    sheet['J1'] = 'GBV'
    sheet['K1'] = 'NBV'
    # sheet['L1'] = 'currency'
    # sheet['M1'] = 'OKOF_OLD'
    # sheet['N1'] = 'CLIENT_CL_NAME'
    # sheet['O1'] = 'STATUS'
    # sheet['P1'] = 'GBV'
    # sheet['Q1'] = 'ACC_DEP'
    # sheet['R1'] = 'NBV'
    # sheet['S1'] = 'reserved'

    for r_num, row in enumerate(results):
        for cell_num, cell in enumerate(row):
            sheet.cell(r_num + 2, cell_num + 1).value = cell

    wb.save(filename=save_path)


if __name__ == "__main__":
    ws = process_xlsx(full_data_path)
    save_results_to_xlsx(ws, full_save_path)
