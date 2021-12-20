import openpyxl as opx
import datetime
from pathlib import Path
import config

open_path = input('Enter path to file:')
file_to_process = input('Enter filename:')
result_file = file_to_process[:-5] + '_to_db_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + '.xlsx'
result_file_1 = file_to_process[:-5] + '_to_db_' + datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S') + '.txt'

full_data_path = Path(open_path).joinpath(file_to_process)
full_save_path = Path(open_path).joinpath(result_file)
full_save_path_1 = Path(open_path).joinpath(result_file_1)
full_save_path_2 = Path(open_path).joinpath(result_file)

def process_xlsx(data_open_path, data_workbook_num=0):
    wb = opx.load_workbook(filename=data_open_path)

    """Default location of data to be processed - firts sheet of Excel workbook"""

    sheet = wb.worksheets[data_workbook_num]
    print('Xlsx data read completed\n')

    row_num = sheet.max_row
    col_num = sheet.max_column

    data_l = [[None for t in range(config.original_columns_num)] for k in range(row_num)]

    """Veriables below to be changed dependant on the structure of the data processed"""

    # print(data_l)
    client_class_name_prev = None
    subdiv_name_prev = None
    # name_prev = None

    for r_num in range(1, row_num+1):

        data_l[r_num - 1][0] = config.div
        if sheet.cell(r_num, 2).value is None:
            subdiv_name_prev = str(sheet.cell(r_num, 1).value).strip()
            data_l[r_num - 1][1] = str(sheet.cell(r_num, 1).value).strip()
            client_class_name_prev = None
            # name_prev = None
            # inv_prev = None
            # mol_name_prev_prev = None
            # date_in_prev = None
            # okof_prev = None
            # status = None
            # gbv_prev = None
            # nbv_prev = None
            # comments_prev = None
        else:
            data_l[r_num - 1][1] = subdiv_name_prev

        if sheet.cell(r_num, 2).value is not None and sheet.cell(r_num, 3).value is None:
            client_class_name_prev = str(sheet.cell(r_num, 1).value).strip()
            data_l[r_num - 1][2] = str(sheet.cell(r_num, 1).value).strip()
            # name_prev = None
            # inv_prev = None
            # mol_name_prev_prev = None
            # date_in_prev = None
            # okof_prev = None
            # status = None
            # gbv_prev = None
            # nbv_prev = None
            # comments_prev = None
        else:
            data_l[r_num - 1][2] = client_class_name_prev

        if sheet.cell(r_num, 2).value is not None and sheet.cell(r_num, 2).value is not None:
            data_l[r_num - 1][3] = sheet.cell(r_num, 1).value

            data_l[r_num - 1][4] = str(sheet.cell(r_num, 2).value).strip() # inv
            data_l[r_num - 1][5] = str(sheet.cell(r_num, 3).value).strip()  # OKOF
            data_l[r_num - 1][6] = str(sheet.cell(r_num, 6).value).strip()  # subdiv_name
            data_l[r_num - 1][7] = str(sheet.cell(r_num, 7).value).strip()  # subdiv
            data_l[r_num - 1][8] = str(sheet.cell(r_num, 8).value).strip()  # HC
            data_l[r_num - 1][9] = str(sheet.cell(r_num, 9).value).strip()  # date_in
            data_l[r_num - 1][10] = str(sheet.cell(r_num, 10).value).strip()  # HC_sign
            data_l[r_num - 1][11] = str(sheet.cell(r_num, 11).value).strip()  # currency
            data_l[r_num - 1][12] = str(sheet.cell(r_num, 12).value).strip()  # OKOF_OLD
            data_l[r_num - 1][13] = str(sheet.cell(r_num, 13).value).strip()  # CL_CLASS_NAME
            data_l[r_num - 1][14] = str(sheet.cell(r_num, 14).value).strip()  # status
            data_l[r_num - 1][15] = str(sheet.cell(r_num, 15).value).strip()  # GBV
            data_l[r_num - 1][16] = str(sheet.cell(r_num, 16).value).strip()  # acc_dep
            data_l[r_num - 1][17] = str(sheet.cell(r_num, 17).value).strip()  # NBV

    return data_l
